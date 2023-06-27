import openpyxl
class Graph:
  #Construtor do grafo
  def __init__(self, num_vert = 0, lista_adj = None, mat_adj = None , arestas = None):
    self.num_vert = num_vert
    if lista_adj == None:
      self.lista_adj = [[]for _ in range(num_vert)]
    else: 
      self.lista_adj = lista_adj
    if arestas == None:
      self.arestas = [[]for _ in range(num_vert)]
    else:
      self.arestas = arestas
    
  #Adicionar aresta com valor 1 do vértice u ao vértice v
  def add_aresta(self, u, v, w = 1):
    self.num_arestas += 1
    if u < self.num_vert and v < self.num_vert:
      self.arestas.append((u, v, w))
      self.lista_adj[u].append((v, w)) 
    else:
      print("Aresta inválida!")
      
  #Remove uma aresta
  def remove_aresta(self, u, v):
    if u < self.num_vert and v < self.num_vert:
      if self.mat_adj[u][v] != 0:
        self.num_arestas += 1
        self.mat_adj[u][v] = 0
        for (v2, w2) in self.lista_adj[u]:
          if v2 == v:
            self.lista_adj[u].remove((v2, w2))
            break
      else:
        print("Aresta inexistente!")
    else:
      print("Aresta invalida!")
 
  #Ler um arquivo xlsx 
  def ler_arquivo(self, nome_arq):
      
    try:
        workbook = openpyxl.load_workbook(r"C:\Users\Ronaldo\Documents\AEDS3\TP01\\"+nome_arq)
        sheet = workbook.active

        # Leitura do cabeçalho
        coluna1 = sheet.cell(row=1, column=1).value
        if self.num_vert > 1000:
            self.repre = True

        coluna2 = sheet.cell(row=1, column=2).value

        # Inicializacao das estruturas de dados
        self.lista_adj = [[] for _ in range(int(self.num_vert))]
        self.mat_adj = [[0 for _ in range(self.num_vert)] for _ in range(self.num_vert)]

        # Le cada aresta do arquivo
        for i in range(2, cont_arestas + 2):
            u = sheet.cell(row=i, column=1).value
            v = sheet.cell(row=i, column=2).value
            w = sheet.cell(row=i, column=3).value
            self.add_aresta(u, v, w)

            if w == 0:
                self.pond = False

            if w < 0:
                self.negativo = True
    except FileNotFoundError:
        print("Nao foi possivel encontrar ou ler o arquivo!")


 